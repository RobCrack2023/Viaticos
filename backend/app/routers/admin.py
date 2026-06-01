from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from typing import List, Optional
import os
from ..database import get_db
from ..models.user import User
from ..models.client import Client
from ..models.project import Project
from ..models.action_type import ActionType
from ..core.config import settings
from ..models.viatico import Viatico, ViaticoMovement
from ..schemas.user import UserCreate, UserUpdate, UserOut
from ..schemas.client import ClientCreate, ClientUpdate, ClientOut
from ..schemas.project import ProjectCreate, ProjectUpdate, ProjectOut
from ..schemas.action_type import ActionTypeCreate, ActionTypeUpdate, ActionTypeOut
from ..schemas.viatico import ViaticoOut
from ..core.security import hash_password
from .auth import require_admin

router = APIRouter(prefix="/admin", tags=["admin"])


# ── Viáticos (todos los usuarios) ─────────────────────────────────────────────

def _build_viatico_out(v: Viatico) -> dict:
    total = sum(m.monto for m in v.movements)
    saldo = v.monto_asignado - total
    return {
        "id": v.id,
        "user_id": v.user_id,
        "user_nombre": v.user.nombre if v.user else "",
        "client_id": v.client_id,
        "project_id": v.project_id,
        "action_type_id": v.action_type_id,
        "client_nombre": v.client.nombre if v.client else "",
        "project_nombre": v.project.nombre if v.project else "",
        "action_type_nombre": v.action_type.nombre if v.action_type else "",
        "monto_asignado": v.monto_asignado,
        "status": v.status,
        "fecha_inicio": v.fecha_inicio,
        "fecha_cierre": v.fecha_cierre,
        "observaciones": v.observaciones,
        "saldo_actual": saldo,
        "total_gastos": total,
        "movements": v.movements,
    }


@router.get("/export/excel")
def export_all_excel(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    from datetime import datetime as dt
    viaticos = db.query(Viatico).order_by(Viatico.fecha_inicio.desc()).all()

    wb = Workbook()
    # ── Hoja resumen ──────────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Resumen"
    blue = "2563EB"
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor=blue)
    thin = Border(left=Side(style="thin",color="E5E7EB"),right=Side(style="thin",color="E5E7EB"),
                  top=Side(style="thin",color="E5E7EB"),bottom=Side(style="thin",color="E5E7EB"))
    for w,col in zip([5,20,20,20,15,14,14,14,12],range(1,10)): ws.column_dimensions[chr(64+col)].width = w
    ws.merge_cells("A1:I1")
    ws["A1"] = f"RESUMEN GENERAL DE VIATICOS — Generado {dt.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = Font(bold=True, size=13, color=blue); ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    headers = ["#","Usuario","Cliente","Proyecto","Tipo Accion","Inicio","Cierre","Asignado","Gastado","Saldo","Estado"]
    ws.column_dimensions["J"].width = 14; ws.column_dimensions["K"].width = 12
    for ci,h in enumerate(headers,1):
        c = ws.cell(row=3,column=ci,value=h); c.font=hf; c.fill=hfill; c.alignment=Alignment(horizontal="center"); c.border=thin

    total_asig = total_gast = 0
    for i,v in enumerate(viaticos,1):
        gastos = sum(m.monto for m in v.movements)
        saldo  = v.monto_asignado - gastos
        total_asig += v.monto_asignado; total_gast += gastos
        row = 3+i
        fill = PatternFill("solid",fgColor="F9FAFB") if i%2==0 else None
        vals = [i, v.user.nombre if v.user else "?",
                v.client.nombre if v.client else "?",
                v.project.nombre if v.project else "?",
                v.action_type.nombre if v.action_type else "?",
                v.fecha_inicio.strftime("%d/%m/%Y") if v.fecha_inicio else "-",
                v.fecha_cierre.strftime("%d/%m/%Y") if v.fecha_cierre else "-",
                v.monto_asignado, gastos, saldo, v.status.value.upper()]
        for ci,val in enumerate(vals,1):
            c = ws.cell(row=row,column=ci,value=val); c.border=thin
            if fill: c.fill=fill
            if ci in (8,9,10): c.number_format='#,##0'; c.alignment=Alignment(horizontal="right")
            if ci==10 and val<0: c.font=Font(color="DC2626",bold=True)
            elif ci==10 and val>=0: c.font=Font(color="059669")

    # Totales
    tr = 3+len(viaticos)+1
    ws.cell(row=tr,column=7,value="TOTALES").font = Font(bold=True)
    for ci,val in [(8,total_asig),(9,total_gast),(10,total_asig-total_gast)]:
        c = ws.cell(row=tr,column=ci,value=val)
        c.font=Font(bold=True); c.number_format='#,##0'; c.alignment=Alignment(horizontal="right")

    # ── Una hoja por usuario ──────────────────────────────────────────────────
    users_viat = {}
    for v in viaticos:
        uid = v.user_id
        users_viat.setdefault(uid, {"nombre": v.user.nombre if v.user else "?", "viaticos": []})
        users_viat[uid]["viaticos"].append(v)

    for uid, ud in users_viat.items():
        nombre = ud["nombre"][:28]
        wsu = wb.create_sheet(nombre)
        wsu.column_dimensions["A"].width = 5; wsu.column_dimensions["B"].width = 14
        wsu.column_dimensions["C"].width = 14; wsu.column_dimensions["D"].width = 30
        wsu.column_dimensions["E"].width = 14; wsu.column_dimensions["F"].width = 14
        wsu.merge_cells("A1:F1")
        wsu["A1"] = f"Viaticos de {ud['nombre']}"
        wsu["A1"].font = Font(bold=True, size=12, color=blue); wsu["A1"].alignment = Alignment(horizontal="center")
        for ci,h in enumerate(["#","Tipo","Fecha","Concepto","Monto","Categoria"],1):
            c = wsu.cell(row=3,column=ci,value=h); c.font=hf; c.fill=hfill; c.border=thin
        row = 4
        for v in ud["viaticos"]:
            wsu.cell(row=row,column=1,value=f"VIATICO #{v.id}: {v.project.nombre if v.project else '?'} ({v.status.value.upper()})").font=Font(bold=True,color=blue)
            wsu.merge_cells(f"A{row}:F{row}"); row+=1
            for i,m in enumerate(v.movements,1):
                fill = PatternFill("solid",fgColor="F9FAFB") if i%2==0 else None
                for ci,val in enumerate([i,m.tipo.upper(),m.fecha.strftime("%d/%m/%Y"),m.concepto,m.monto,getattr(m,'categoria','')],1):
                    c = wsu.cell(row=row,column=ci,value=val); c.border=thin
                    if fill: c.fill=fill
                    if ci==5: c.number_format='#,##0'; c.alignment=Alignment(horizontal="right")
                row+=1
            row+=1

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"viaticos_completo_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/stats")
def get_stats(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    viaticos = db.query(Viatico).all()

    by_user, by_project, by_action = {}, {}, {}

    for v in viaticos:
        gastos = sum(m.monto for m in v.movements)
        is_closed = v.status.value == "cerrado"

        # Por usuario
        uid = v.user_id
        if uid not in by_user:
            by_user[uid] = {"nombre": v.user.nombre if v.user else "?", "total": 0, "activos": 0, "cerrados": 0, "asignado": 0.0, "gastado": 0.0}
        by_user[uid]["total"] += 1
        by_user[uid]["activos" if not is_closed else "cerrados"] += 1
        by_user[uid]["asignado"] += v.monto_asignado
        by_user[uid]["gastado"] += gastos

        # Por proyecto
        pid = v.project_id
        if pid not in by_project:
            by_project[pid] = {"nombre": v.project.nombre if v.project else "?", "cliente": v.client.nombre if v.client else "?", "total": 0, "asignado": 0.0, "gastado": 0.0}
        by_project[pid]["total"] += 1
        by_project[pid]["asignado"] += v.monto_asignado
        by_project[pid]["gastado"] += gastos

        # Por tipo de acción
        aid = v.action_type_id
        if aid not in by_action:
            by_action[aid] = {"nombre": v.action_type.nombre if v.action_type else "?", "total": 0, "asignado": 0.0, "gastado": 0.0}
        by_action[aid]["total"] += 1
        by_action[aid]["asignado"] += v.monto_asignado
        by_action[aid]["gastado"] += gastos

    total_asignado = sum(v.monto_asignado for v in viaticos)
    total_gastado  = sum(sum(m.monto for m in v.movements) for v in viaticos)

    return {
        "resumen": {
            "total":          len(viaticos),
            "activos":        sum(1 for v in viaticos if v.status.value == "activo"),
            "cerrados":       sum(1 for v in viaticos if v.status.value == "cerrado"),
            "total_asignado": total_asignado,
            "total_gastado":  total_gastado,
            "total_saldo":    total_asignado - total_gastado,
        },
        "por_usuario":     sorted(by_user.values(),    key=lambda x: x["gastado"], reverse=True),
        "por_proyecto":    sorted(by_project.values(), key=lambda x: x["gastado"], reverse=True),
        "por_tipo_accion": sorted(by_action.values(),  key=lambda x: x["gastado"], reverse=True),
    }


@router.delete("/viaticos/{viatico_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_viatico(viatico_id: int, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    v = db.query(Viatico).filter(Viatico.id == viatico_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="Viático no encontrado")

    # 1. Eliminar fotos del sistema de archivos
    for m in v.movements:
        if m.foto_path:
            foto_abs = os.path.join(settings.UPLOADS_DIR, m.foto_path)
            try:
                if os.path.exists(foto_abs):
                    os.remove(foto_abs)
            except Exception:
                pass  # Si falla la eliminación del archivo, continuar igual

    # 2. Eliminar movimientos del viático
    for m in v.movements:
        db.delete(m)

    # 3. Eliminar el viático
    db.delete(v)
    db.commit()


@router.get("/viaticos")
def list_all_viaticos(
    status_filter: Optional[str] = None,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    q = db.query(Viatico)
    if status_filter:
        q = q.filter(Viatico.status == status_filter)
    viaticos = q.order_by(Viatico.fecha_inicio.desc()).all()
    return [_build_viatico_out(v) for v in viaticos]


# ── Usuarios ──────────────────────────────────────────────────────────────────

@router.get("/users", response_model=List[UserOut])
def list_users(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    return db.query(User).all()


@router.post("/users", response_model=UserOut, status_code=status.HTTP_201_CREATED)
def create_user(data: UserCreate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    if db.query(User).filter(User.email == data.email).first():
        raise HTTPException(status_code=400, detail="Email ya registrado")
    user = User(
        nombre=data.nombre,
        email=data.email,
        password_hash=hash_password(data.password),
        is_admin=data.is_admin,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@router.put("/users/{user_id}", response_model=UserOut)
def update_user(user_id: int, data: UserUpdate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if data.nombre is not None:
        user.nombre = data.nombre
    if data.email is not None:
        user.email = data.email
    if data.password is not None:
        user.password_hash = hash_password(data.password)
    if data.is_admin is not None:
        user.is_admin = data.is_admin
    if data.is_active is not None:
        user.is_active = data.is_active
    db.commit()
    db.refresh(user)
    return user


# ── Clientes ──────────────────────────────────────────────────────────────────

@router.get("/clients", response_model=List[ClientOut])
def list_clients(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    return db.query(Client).all()


@router.post("/clients", response_model=ClientOut, status_code=status.HTTP_201_CREATED)
def create_client(data: ClientCreate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    client = Client(**data.model_dump())
    db.add(client)
    db.commit()
    db.refresh(client)
    return client


@router.put("/clients/{client_id}", response_model=ClientOut)
def update_client(client_id: int, data: ClientUpdate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(client, field, value)
    db.commit()
    db.refresh(client)
    return client


# ── Proyectos ─────────────────────────────────────────────────────────────────

@router.get("/projects", response_model=List[ProjectOut])
def list_projects(client_id: int = None, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    q = db.query(Project)
    if client_id:
        q = q.filter(Project.client_id == client_id)
    return q.all()


@router.post("/projects", response_model=ProjectOut, status_code=status.HTTP_201_CREATED)
def create_project(data: ProjectCreate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    if not db.query(Client).filter(Client.id == data.client_id).first():
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    project = Project(**data.model_dump())
    db.add(project)
    db.commit()
    db.refresh(project)
    return project


@router.put("/projects/{project_id}", response_model=ProjectOut)
def update_project(project_id: int, data: ProjectUpdate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(project, field, value)
    db.commit()
    db.refresh(project)
    return project


# ── Tipos de Acción ───────────────────────────────────────────────────────────

@router.get("/action-types", response_model=List[ActionTypeOut])
def list_action_types(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    return db.query(ActionType).all()


@router.post("/action-types", response_model=ActionTypeOut, status_code=status.HTTP_201_CREATED)
def create_action_type(data: ActionTypeCreate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    at = ActionType(**data.model_dump())
    db.add(at)
    db.commit()
    db.refresh(at)
    return at


@router.put("/action-types/{at_id}", response_model=ActionTypeOut)
def update_action_type(at_id: int, data: ActionTypeUpdate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    at = db.query(ActionType).filter(ActionType.id == at_id).first()
    if not at:
        raise HTTPException(status_code=404, detail="Tipo de acción no encontrado")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(at, field, value)
    db.commit()
    db.refresh(at)
    return at
