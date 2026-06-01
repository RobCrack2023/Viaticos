from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
import os
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable, KeepTogether
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..database import get_db
from ..models.user import User
from ..models.viatico import Viatico, ViaticoStatus
from ..models.account import Account
from .auth import get_current_user
from ..core.config import settings

router = APIRouter(prefix="/reports", tags=["reportes"])

CLP = lambda v: f"${v:,.0f}".replace(",", ".")


def _get_viatico(viatico_id: int, user: User, db: Session) -> Viatico:
    q = db.query(Viatico).filter(Viatico.id == viatico_id)
    # Admin puede ver cualquier viático; usuario solo los suyos
    if not user.is_admin:
        q = q.filter(Viatico.user_id == user.id)
    v = q.first()
    if not v:
        raise HTTPException(status_code=404, detail="Viático no encontrado")
    return v


def _calc(v: Viatico):
    total = sum(m.monto for m in v.movements)
    saldo = v.monto_asignado - total
    return total, saldo


# ── PDF ───────────────────────────────────────────────────────────────────────

@router.get("/{viatico_id}/pdf")
def download_pdf(viatico_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    v = _get_viatico(viatico_id, current_user, db)
    total_gastos, saldo = _calc(v)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, alignment=TA_CENTER, spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, textColor=colors.grey, alignment=TA_CENTER)
    label_style = ParagraphStyle("label", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    value_style = ParagraphStyle("value", parent=styles["Normal"], fontSize=10, fontName="Helvetica-Bold")
    right_style = ParagraphStyle("right", parent=styles["Normal"], fontSize=10, alignment=TA_RIGHT)

    story.append(Paragraph("RENDICIÓN DE VIÁTICO", title_style))
    story.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2563EB")))
    story.append(Spacer(1, 0.4*cm))

    # Encabezado de información
    info_data = [
        ["Colaborador:", current_user.nombre, "Estado:", v.status.value.upper()],
        ["Cliente:", v.client.nombre, "Fecha inicio:", v.fecha_inicio.strftime("%d/%m/%Y")],
        ["Proyecto:", v.project.nombre, "Fecha cierre:", v.fecha_cierre.strftime("%d/%m/%Y") if v.fecha_cierre else "-"],
        ["Tipo de acción:", v.action_type.nombre, "", ""],
    ]
    info_table = Table(info_data, colWidths=[3.5*cm, 7*cm, 3.5*cm, 3*cm])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6B7280")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#6B7280")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.4*cm))

    if v.observaciones:
        story.append(Paragraph(f"Observaciones: {v.observaciones}", label_style))
        story.append(Spacer(1, 0.3*cm))

    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph("DETALLE DE MOVIMIENTOS", ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11, spaceAfter=6)))

    headers = ["#", "Fecha", "Tipo", "Concepto", "Monto"]
    rows = [headers]
    for i, m in enumerate(v.movements, 1):
        rows.append([
            str(i),
            m.fecha.strftime("%d/%m/%Y"),
            m.tipo.upper(),
            m.concepto,
            CLP(m.monto),
        ])

    col_w = [0.8*cm, 2.5*cm, 2*cm, 9*cm, 3*cm]
    mv_table = Table(rows, colWidths=col_w)
    mv_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(mv_table)
    story.append(Spacer(1, 0.6*cm))

    # Resumen financiero
    devolucion = saldo >= 0
    color_resultado = colors.HexColor("#16A34A") if devolucion else colors.HexColor("#DC2626")
    label_resultado = "DEBE DEVOLVER" if devolucion else "A REEMBOLSAR"

    resumen_data = [
        ["Monto asignado:", CLP(v.monto_asignado)],
        ["Total gastos:", CLP(total_gastos)],
        [label_resultado + ":", CLP(abs(saldo))],
    ]
    resumen_table = Table(resumen_data, colWidths=[5*cm, 3*cm], hAlign="RIGHT")
    resumen_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -2), "Helvetica"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TEXTCOLOR", (0, -1), (-1, -1), color_resultado),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#E5E7EB")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(resumen_table)

    # ── Sección de fotos de comprobantes ────────────────────────────────────────
    movs_con_foto = [(i, m) for i, m in enumerate(v.movements, 1) if m.foto_path]
    if movs_con_foto:
        story.append(Spacer(1, 0.8*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2563EB")))
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph("COMPROBANTES FOTOGRÁFICOS", ParagraphStyle("h2foto", parent=styles["Heading2"], fontSize=11, spaceAfter=8)))

        MAX_W = 15 * cm
        MAX_H = 10 * cm

        for i, m in movs_con_foto:
            foto_abs = os.path.join(settings.UPLOADS_DIR, m.foto_path)
            if not os.path.exists(foto_abs):
                continue
            try:
                # Calcular dimensiones proporcionales con Pillow
                with PILImage.open(foto_abs) as pil:
                    pw, ph = pil.size
                ratio = min(MAX_W / pw, MAX_H / ph)
                img_w, img_h = pw * ratio, ph * ratio

                label = ParagraphStyle("foto_label", parent=styles["Normal"], fontSize=9,
                                       textColor=colors.HexColor("#6B7280"), spaceBefore=4, spaceAfter=4)
                bloque = [
                    Paragraph(f"<b>#{i}</b> — {m.tipo.upper()} | {m.concepto} | {m.fecha.strftime('%d/%m/%Y')} | {CLP(m.monto)}", label),
                    Image(foto_abs, width=img_w, height=img_h),
                    Spacer(1, 0.5*cm),
                ]
                story.append(KeepTogether(bloque))
            except Exception:
                story.append(Paragraph(f"#{i} — {m.concepto}: imagen no disponible",
                                       ParagraphStyle("err", parent=styles["Normal"], fontSize=8, textColor=colors.grey)))

    doc.build(story)
    buf.seek(0)
    filename = f"rendicion_viatico_{viatico_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


# ── Excel ─────────────────────────────────────────────────────────────────────

@router.get("/{viatico_id}/excel")
def download_excel(viatico_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    v = _get_viatico(viatico_id, current_user, db)
    total_gastos, saldo = _calc(v)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rendición"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 16

    blue = "2563EB"
    light = "F9FAFB"
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor=blue)
    bold = Font(bold=True, size=10)
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    # Título
    ws.merge_cells("A1:E1")
    ws["A1"] = "RENDICIÓN DE VIÁTICO"
    ws["A1"].font = Font(bold=True, size=14, color=blue)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    # Info
    info = [
        ("Colaborador:", current_user.nombre),
        ("Cliente:", v.client.nombre),
        ("Proyecto:", v.project.nombre),
        ("Tipo de acción:", v.action_type.nombre),
        ("Fecha inicio:", v.fecha_inicio.strftime("%d/%m/%Y")),
        ("Fecha cierre:", v.fecha_cierre.strftime("%d/%m/%Y") if v.fecha_cierre else "-"),
        ("Estado:", v.status.value.upper()),
    ]
    for i, (lbl, val) in enumerate(info, 3):
        ws[f"A{i}"] = lbl
        ws[f"A{i}"].font = Font(bold=True, color="6B7280", size=9)
        ws[f"B{i}"] = val
        ws[f"B{i}"].font = Font(size=10)

    # Tabla movimientos
    start = 12
    headers = ["#", "Fecha", "Tipo", "Concepto", "Monto"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for i, m in enumerate(v.movements, 1):
        row = start + i
        fill = PatternFill("solid", fgColor=light) if i % 2 == 0 else None
        values = [i, m.fecha.strftime("%d/%m/%Y"), m.tipo.upper(), m.concepto, m.monto]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            if fill:
                cell.fill = fill
            if col == 5:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

    # Resumen
    sum_row = start + len(v.movements) + 2
    resumen = [
        ("Monto asignado:", v.monto_asignado),
        ("Total gastos:", total_gastos),
        ("DEBE DEVOLVER:" if saldo >= 0 else "A REEMBOLSAR:", abs(saldo)),
    ]
    for j, (lbl, val) in enumerate(resumen):
        r = sum_row + j
        ws[f"D{r}"] = lbl
        ws[f"D{r}"].font = bold if j == 2 else Font(size=10)
        ws[f"E{r}"] = val
        ws[f"E{r}"].number_format = '#,##0'
        ws[f"E{r}"].alignment = Alignment(horizontal="right")
        if j == 2:
            color = "16A34A" if saldo >= 0 else "DC2626"
            ws[f"D{r}"].font = Font(bold=True, color=color, size=10)
            ws[f"E{r}"].font = Font(bold=True, color=color, size=10)

    # ── Hoja 2: Fotos de comprobantes ─────────────────────────────────────────
    movs_con_foto = [(i, m) for i, m in enumerate(v.movements, 1) if m.foto_path]
    if movs_con_foto:
        from openpyxl.drawing.image import Image as XLImage
        ws2 = wb.create_sheet("Comprobantes")
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 60

        ws2["A1"] = "COMPROBANTES FOTOGRÁFICOS"
        ws2["A1"].font = Font(bold=True, size=13, color=blue)
        ws2.merge_cells("A1:B1")
        ws2.row_dimensions[1].height = 22

        cur_row = 3
        for i, m in movs_con_foto:
            foto_abs = os.path.join(settings.UPLOADS_DIR, m.foto_path)
            if not os.path.exists(foto_abs):
                continue
            try:
                # Etiqueta del movimiento
                ws2[f"A{cur_row}"] = f"#{i}"
                ws2[f"A{cur_row}"].font = Font(bold=True, size=10, color=blue)
                ws2[f"B{cur_row}"] = f"{m.tipo.upper()} | {m.concepto} | {m.fecha.strftime('%d/%m/%Y')} | ${m.monto:,.0f}"
                ws2[f"B{cur_row}"].font = Font(size=10)
                cur_row += 1

                # Calcular tamaño de imagen (max 500px de ancho)
                with PILImage.open(foto_abs) as pil:
                    pw, ph = pil.size
                max_px = 500
                ratio = min(max_px / pw, max_px / ph, 1)
                img_w = int(pw * ratio)
                img_h = int(ph * ratio)

                xl_img = XLImage(foto_abs)
                xl_img.width  = img_w
                xl_img.height = img_h
                ws2.add_image(xl_img, f"B{cur_row}")

                # Ajustar altura de filas para la imagen
                rows_needed = max(1, img_h // 15)
                for r in range(cur_row, cur_row + rows_needed):
                    ws2.row_dimensions[r].height = 15
                cur_row += rows_needed + 2
            except Exception:
                ws2[f"B{cur_row}"] = f"#{i} — imagen no disponible"
                cur_row += 2

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"rendicion_viatico_{viatico_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


# ── Exportar Cuenta Corriente ─────────────────────────────────────────────────

def _get_account_report(user: User, db: Session) -> Account:
    acc = db.query(Account).filter(Account.user_id == user.id).first()
    if not acc:
        raise HTTPException(status_code=404, detail="Cuenta corriente no inicializada")
    return acc


def _calc_account_saldo(acc: Account) -> float:
    s = acc.saldo_inicial
    for m in acc.movements:
        s += m.monto if m.tipo == "ingreso" else -m.monto
    return s


@router.get("/cc/pdf")
def account_pdf(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    acc = _get_account_report(current_user, db)
    saldo = _calc_account_saldo(acc)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle("t", parent=styles["Heading1"], fontSize=16, alignment=TA_CENTER, spaceAfter=6)
    sub_style   = ParagraphStyle("s", parent=styles["Normal"],   fontSize=10, textColor=colors.grey, alignment=TA_CENTER)
    label_style = ParagraphStyle("l", parent=styles["Normal"],   fontSize=9,  textColor=colors.grey)

    story.append(Paragraph("ESTADO DE CUENTA CORRIENTE", title_style))
    story.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2563EB")))
    story.append(Spacer(1, 0.4*cm))

    info_data = [
        ["Colaborador:", current_user.nombre],
        ["Saldo inicial:", CLP(acc.saldo_inicial)],
        ["Saldo actual:", CLP(saldo)],
    ]
    it = Table(info_data, colWidths=[4*cm, 10*cm])
    it.setStyle(TableStyle([("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),
                             ("TEXTCOLOR",(0,0),(0,-1),colors.HexColor("#6B7280")),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
    story.append(it)
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph("MOVIMIENTOS", ParagraphStyle("h2",parent=styles["Heading2"],fontSize=11,spaceAfter=6)))

    headers = ["#", "Fecha", "Tipo", "N° Doc", "Concepto", "Monto"]
    rows = [headers]
    for i, m in enumerate(acc.movements, 1):
        rows.append([str(i), m.fecha.strftime("%d/%m/%Y"), m.tipo.upper(),
                     m.numero_doc or "-", m.concepto, CLP(m.monto)])

    mv_table = Table(rows, colWidths=[0.6*cm, 2.2*cm, 1.8*cm, 2.2*cm, 8*cm, 2.5*cm])
    mv_table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2563EB")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8),("ALIGN",(5,0),(5,-1),"RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F9FAFB")]),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E5E7EB")),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),("TOPPADDING",(0,0),(-1,-1),5),
    ]))
    story.append(mv_table)
    story.append(Spacer(1, 0.6*cm))

    res_data = [["Saldo inicial:", CLP(acc.saldo_inicial)],
                ["Total ingresos:", CLP(sum(m.monto for m in acc.movements if m.tipo=="ingreso"))],
                ["Total egresos:", CLP(sum(m.monto for m in acc.movements if m.tipo!="ingreso"))],
                ["SALDO ACTUAL:", CLP(saldo)]]
    rt = Table(res_data, colWidths=[5*cm, 3*cm], hAlign="RIGHT")
    rt.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,-2),"Helvetica"),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),10),("ALIGN",(1,0),(1,-1),"RIGHT"),
        ("LINEABOVE",(0,-1),(-1,-1),1,colors.HexColor("#E5E7EB")),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
    ]))
    story.append(rt)

    doc.build(story)
    buf.seek(0)
    filename = f"cuenta_corriente_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/cc/excel")
def account_excel(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    acc = _get_account_report(current_user, db)
    saldo = _calc_account_saldo(acc)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuenta Corriente"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 16
    blue = "2563EB"
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor=blue)
    thin = Border(left=Side(style="thin",color="E5E7EB"),right=Side(style="thin",color="E5E7EB"),
                  top=Side(style="thin",color="E5E7EB"),bottom=Side(style="thin",color="E5E7EB"))

    ws.merge_cells("A1:F1")
    ws["A1"] = f"ESTADO DE CUENTA CORRIENTE — {current_user.nombre}"
    ws["A1"].font = Font(bold=True, size=13, color=blue)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    for i, (lbl, val) in enumerate([("Saldo inicial:",CLP(acc.saldo_inicial)),("Saldo actual:",CLP(saldo))], 3):
        ws[f"A{i}"] = lbl; ws[f"A{i}"].font = Font(bold=True, color="6B7280", size=9)
        ws[f"B{i}"] = val; ws[f"B{i}"].font = Font(size=10)

    start = 7
    for ci, h in enumerate(["#","Fecha","Tipo","N° Doc","Concepto","Monto"], 1):
        c = ws.cell(row=start, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center"); c.border = thin

    light = "F9FAFB"
    for i, m in enumerate(acc.movements, 1):
        row = start + i
        fill = PatternFill("solid", fgColor=light) if i%2==0 else None
        for ci, val in enumerate([i, m.fecha.strftime("%d/%m/%Y"), m.tipo.upper(),
                                   m.numero_doc or "-", m.concepto, m.monto], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = thin
            if fill: c.fill = fill
            if ci == 6: c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"cuenta_corriente_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})
